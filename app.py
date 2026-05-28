import io
import os
from datetime import datetime

import folium
import geopandas as gpd
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from shapely.geometry import Point
from streamlit_folium import st_folium
from streamlit_autorefresh import st_autorefresh

# ── Auto-refresh every 2 minutes ─────────────────────────────────────
st_autorefresh(interval=120000, key="data_refresh")

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SHEET_ID = "1seIjUlWoN3soQKgr2u8Hd7vTXDnqYZcCoDYwW1x38tY"
SHEET_CSV_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv"
)
GEOJSON_PATH = os.path.join(os.path.dirname(__file__), "assets", "multan_district.geojson")

# One distinct colour per Tehsil
TEHSIL_PALETTE = {
    "Multan-1":            "#4285F4",
    "Multan-2":            "#EA4335",
    "Multan-3":            "#FBBC04",
    "Shujabad":            "#34A853",
    "Jalalpur Pir Wala":   "#FF6D00",
}

def norm_status(raw):
    """Normalise status string to canonical form."""
    s = str(raw).strip().lower()
    if "resolv" in s:
        return "Resolved"
    if "pend" in s:
        return "Pending"
    if "process" in s or "progress" in s:
        return "In Process"
    if "clos" in s:
        return "Closed"
    return str(raw).strip()

STATUS_MARKER_COLOR = {
    "Pending":    "red",
    "Resolved":   "green",
    "In Process": "#FFD700",
    "Closed":     "gray",
}

STATUS_LABEL_COLOR = {
    "Pending":    "#E53935",
    "Resolved":   "#43A047",
    "In Process": "#D4A800",
    "Closed":     "#757575",
}

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG & GLOBAL CSS
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SPA Multan – Special Branch Complaints Dashboard",
    layout="wide",
    page_icon="📋",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
        /* ── header bar ── */
        .dash-header {
            background: linear-gradient(135deg, #1a6b3b 0%, #0d4a28 100%);
            border-radius: 12px;
            padding: 18px 28px;
            margin-bottom: 18px;
            color: white;
        }
        .dash-header h1 { color: white; margin: 0; font-size: 26px; }
        .dash-header p  { color: #c8f0d8; margin: 4px 0 0; font-size: 13px; }

        /* ── metric cards ── */
        .metric-row { display: flex; gap: 12px; margin-bottom: 16px; }
        .metric-card {
            flex: 1;
            background: white;
            border-radius: 10px;
            padding: 14px 18px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            border-left: 5px solid;
        }
        .metric-card .val { font-size: 32px; font-weight: 700; line-height: 1; }
        .metric-card .lbl { font-size: 12px; color: #777; margin-top: 4px; }

        /* ── section titles ── */
        .section-title {
            font-size: 16px;
            font-weight: 600;
            color: #1a6b3b;
            border-bottom: 2px solid #e0f2e9;
            padding-bottom: 6px;
            margin-bottom: 10px;
        }

        /* legend pill */
        .legend-pill {
            display: inline-block;
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            margin: 3px;
            color: white;
        }

        /* ── metric label bold, value normal size ── */
        [data-testid="stMetricLabel"] p {
            font-size: 15px !important;
            font-weight: 700 !important;
            color: #1a3a2a !important;
        }
        [data-testid="stMetricValue"] {
            font-size: 28px !important;
            font-weight: 700 !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADERS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=120, show_spinner=False)
def load_sheet_data() -> pd.DataFrame:
    """Fetch complaints from Google Sheets CSV export."""
    resp = requests.get(SHEET_CSV_URL, timeout=20, allow_redirects=True)
    resp.raise_for_status()
    df = pd.read_csv(io.StringIO(resp.text))
    df.columns = df.columns.str.strip()

    # Normalise Status column
    if "Status" in df.columns:
        df["Status"] = df["Status"].apply(norm_status)

    # Parse GPS → separate Latitude / Longitude columns
    def parse_gps(val):
        try:
            lat, lon = str(val).split(",")
            return float(lat.strip()), float(lon.strip())
        except Exception:
            return None, None

    df[["Latitude", "Longitude"]] = df["GPS"].apply(
        lambda v: pd.Series(parse_gps(v))
    )

    # Keep only rows with valid coordinates
    df = df.dropna(subset=["Latitude", "Longitude"]).reset_index(drop=True)

    # Parse date
    df["Date"] = pd.to_datetime(df.get("Date"), errors="coerce")
    return df


@st.cache_resource(show_spinner=False)
def load_geojson() -> gpd.GeoDataFrame:
    gdf = gpd.read_file(GEOJSON_PATH)
    if gdf.crs is None:
        gdf = gdf.set_crs("EPSG:4326")
    else:
        gdf = gdf.to_crs("EPSG:4326")
    return gdf


def assign_uc_tehsil(df: pd.DataFrame, gdf: gpd.GeoDataFrame) -> pd.DataFrame:
    """Spatial join: tag each complaint with the UC it falls inside."""
    if df.empty:
        df["UC_Code"]   = pd.NA
        df["UC_No"]     = pd.NA
        df["UC_Tehsil"] = pd.NA
        return df

    pts = gpd.GeoDataFrame(
        df.copy(),
        geometry=[Point(row.Longitude, row.Latitude) for _, row in df.iterrows()],
        crs="EPSG:4326",
    )
    uc_cols = gdf[["UC_Code", "UC_No", "Tehsil_N", "geometry"]].copy()
    joined = gpd.sjoin(pts, uc_cols, how="left", predicate="within")

    # sjoin may produce duplicates when a point sits on a shared boundary
    joined = joined[~joined.index.duplicated(keep="first")]

    df = df.copy()
    df["UC_Code"]   = joined["UC_Code"].values
    df["UC_No"]     = joined["UC_No"].values
    df["UC_Tehsil"] = joined["Tehsil_N"].values
    return df


# ─────────────────────────────────────────────────────────────────────────────
# MAP BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def build_map(df_filtered: pd.DataFrame, gdf: gpd.GeoDataFrame) -> folium.Map:
    center_lat = (gdf.total_bounds[1] + gdf.total_bounds[3]) / 2
    center_lon = (gdf.total_bounds[0] + gdf.total_bounds[2]) / 2

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=10,
        tiles="OpenStreetMap",
    )

    # ── UC Boundaries coloured by Tehsil ──────────────────────────────────
    def uc_style(feature):
        tehsil = feature["properties"].get("Tehsil_N", "")
        return {
            "fillColor":   TEHSIL_PALETTE.get(tehsil, "#AAAAAA"),
            "color":       "#444",
            "weight":      1.0,
            "fillOpacity": 0.28,
        }

    folium.GeoJson(
        gdf.__geo_interface__,
        name="UC Boundaries",
        style_function=uc_style,
        tooltip=folium.GeoJsonTooltip(
            fields=["UC_Code", "Tehsil_N", "UC_No"],
            aliases=["UC Code", "Tehsil", "UC No"],
            sticky=True,
            style="font-size:13px;",
        ),
    ).add_to(m)

    # ── Complaint Markers ─────────────────────────────────────────────────
    complaints_group = folium.FeatureGroup(name="Complaints", show=True)
    for _, row in df_filtered.iterrows():
        status     = norm_status(row.get("Status", "Pending"))
        mcolor     = STATUS_MARKER_COLOR.get(status, "red")
        lcolor     = STATUS_LABEL_COLOR.get(status, "#333333")
        app_no     = row.get("Application Number", "")
        issue      = row.get("Issue", "")
        date_str   = str(row.get("Date", ""))[:10]
        tehsil     = row.get("Tehsil", "")
        address    = row.get("Address", "")
        uc_code    = row.get("UC_Code", "N/A") or "N/A"
        uc_tehsil  = row.get("UC_Tehsil", "N/A") or "N/A"

        popup_html = f"""
        <div style="min-width:230px;font-family:sans-serif;font-size:13px">
          <div style="background:#1a6b3b;color:white;padding:6px 10px;
                      border-radius:6px 6px 0 0;font-weight:700">
            {app_no}
          </div>
          <div style="padding:8px 10px;border:1px solid #ddd;border-top:none;
                      border-radius:0 0 6px 6px">
            <b>Issue:</b> {issue}<br>
            <b>Date:</b> {date_str}<br>
            <b>Tehsil (Sheet):</b> {tehsil}<br>
            <b>Address:</b> {address}<br>
            <b>Status:</b>
              <span style="font-weight:700;color:{lcolor}">{status}</span><br>
            <hr style="margin:6px 0">
            <b>UC Code (GeoJSON):</b> {uc_code}<br>
            <b>Tehsil (GeoJSON):</b> {uc_tehsil}<br>
          </div>
        </div>"""

        folium.CircleMarker(
            location=[row["Latitude"], row["Longitude"]],
            radius=9,
            color=mcolor,
            fill=True,
            fill_color=mcolor,
            fill_opacity=0.85,
            weight=2,
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=f"{app_no} · {status}",
        ).add_to(complaints_group)

    complaints_group.add_to(m)
    folium.LayerControl(collapsed=False).add_to(m)
    return m


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
EXPORT_COLUMNS = [
    "Sr No",
    "Application Number",
    "Date",
    "Issue",
    "Region",
    "District",
    "Tehsil",
    "Address",
    "GPS",
    "Latitude",
    "Longitude",
    "Status",
    "UC_Code",
    "UC_Tehsil",
    "UC_No",
]

FRIENDLY_HEADERS = {
    "UC_Code":   "UC Code",
    "UC_Tehsil": "Tehsil (GeoJSON)",
    "UC_No":     "UC No",
}


def generate_excel(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints"

    # ── styles ─────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="1a6b3b")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    alt_fill  = PatternFill("solid", fgColor="EDF7F1")
    even_fill = PatternFill("solid", fgColor="FFFFFF")
    thin      = Side(border_style="thin", color="D0E8D8")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── header row ─────────────────────────────────────────────────────────
    cols = [c for c in EXPORT_COLUMNS if c in df.columns]
    for ci, col in enumerate(cols, 1):
        cell       = ws.cell(row=1, column=ci, value=FRIENDLY_HEADERS.get(col, col))
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = center
        cell.border = cell_border

    ws.row_dimensions[1].height = 28

    # ── data rows ──────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        row_fill = alt_fill if ri % 2 == 0 else even_fill
        for ci, col in enumerate(cols, 1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            elif col == "Date" and isinstance(val, pd.Timestamp):
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            else:
                val = str(val)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = row_fill
            cell.alignment = left
            cell.border    = cell_border

    # ── column widths ──────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        hdr_len  = len(FRIENDLY_HEADERS.get(col, col))
        data_max = max(
            (len(str(ws.cell(row=r, column=ci).value or ""))
             for r in range(2, ws.max_row + 1)),
            default=0,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(hdr_len, data_max) + 3, 42
        )

    # ── freeze panes & auto-filter ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────────────────────
def main():
    # ── Header ───────────────────────────────────────────────────────────
    st.markdown(
        """
        <div class="dash-header">
          <h1>📋  SPA Multan – Special Branch Complaints Dashboard</h1>
          <p>Live data · Google Sheets &nbsp;|&nbsp; UC Boundaries · Multan District GeoJSON &nbsp;|&nbsp; Base Map · OpenStreetMap</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Load data ─────────────────────────────────────────────────────────
    with st.spinner("Loading complaints from Google Sheets …"):
        try:
            df = load_sheet_data()
        except Exception as err:
            st.error(f"Could not load Google Sheets data: {err}")
            st.stop()

    with st.spinner("Loading UC boundaries …"):
        gdf = load_geojson()

    with st.spinner("Running spatial join …"):
        df = assign_uc_tehsil(df, gdf)

    # ── Sidebar Filters ───────────────────────────────────────────────────
    st.sidebar.markdown("## 🔍 Filters")

    statuses  = ["All"] + sorted(df["Status"].dropna().unique().tolist())
    tehsils   = ["All"] + sorted(df["Tehsil"].dropna().unique().tolist())
    issues    = ["All"] + sorted(df["Issue"].dropna().unique().tolist())
    uc_tehsil_opts = ["All"] + sorted(
        [x for x in df["UC_Tehsil"].dropna().unique().tolist()]
    )

    sel_status    = st.sidebar.selectbox("Status",             statuses)
    sel_tehsil    = st.sidebar.selectbox("Tehsil (Sheet)",     tehsils)
    sel_uc_tehsil = st.sidebar.selectbox("Tehsil (GeoJSON)",   uc_tehsil_opts)
    sel_issue     = st.sidebar.selectbox("Issue Type",         issues)

    # Date range only when valid dates exist
    if df["Date"].notna().any():
        min_d = df["Date"].min().date()
        max_d = df["Date"].max().date()
        date_range = st.sidebar.date_input(
            "Date Range", value=[min_d, max_d], min_value=min_d, max_value=max_d
        )
    else:
        date_range = None

    st.sidebar.divider()
    if st.sidebar.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()

    # ── Tehsil colour legend ──────────────────────────────────────────────
    st.sidebar.markdown("**UC Boundary Colours**")
    legend_html = "".join(
        f'<span class="legend-pill" style="background:{clr}">{teh}</span>'
        for teh, clr in TEHSIL_PALETTE.items()
    )
    st.sidebar.markdown(legend_html, unsafe_allow_html=True)

    # ── Apply Filters ─────────────────────────────────────────────────────
    df_filt = df.copy()
    if sel_status    != "All":
        df_filt = df_filt[df_filt["Status"]    == sel_status]
    if sel_tehsil    != "All":
        df_filt = df_filt[df_filt["Tehsil"]    == sel_tehsil]
    if sel_uc_tehsil != "All":
        df_filt = df_filt[df_filt["UC_Tehsil"] == sel_uc_tehsil]
    if sel_issue     != "All":
        df_filt = df_filt[df_filt["Issue"]     == sel_issue]
    if date_range and len(date_range) == 2:
        df_filt = df_filt[
            (df_filt["Date"].dt.date >= date_range[0])
            & (df_filt["Date"].dt.date <= date_range[1])
        ]

    # ── Metric Cards ──────────────────────────────────────────────────────
    total      = len(df_filt)
    pending    = (df_filt["Status"] == "Pending").sum()
    resolved   = (df_filt["Status"] == "Resolved").sum()
    in_process = (df_filt["Status"] == "In Process").sum()
    unmatched  = df_filt["UC_Code"].isna().sum()

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Complaints",  total)
    c2.metric("Pending",           int(pending))
    c3.metric("Resolved",          int(resolved))
    c4.metric("In Process",        int(in_process))
    c5.metric("Outside UC Bounds", int(unmatched),
              delta=None if unmatched == 0 else "⚠️ Check GPS",
              delta_color="off")

    st.divider()

    # ── Full-width Map ────────────────────────────────────────────────────
    st.markdown('<p class="section-title">🗺️ Complaints Map</p>',
                unsafe_allow_html=True)
    with st.spinner("Rendering map …"):
        fmap = build_map(df_filt, gdf)
    st_folium(fmap, width="100%", height=640, returned_objects=[])

    st.divider()

    # ── Attribute Table (below map) ───────────────────────────────────────
    st.markdown(
        f'<p class="section-title">📄 Complaints Attribute Table &nbsp;'
        f'<span style="color:#888;font-size:13px">({total} shown)</span></p>',
        unsafe_allow_html=True,
    )
    display_cols = [
        "Sr No", "Application Number", "Date", "Issue",
        "Region", "District", "Tehsil", "Address",
        "Status", "UC_Code", "UC_Tehsil", "UC_No",
    ]
    avail = [c for c in display_cols if c in df_filt.columns]
    df_display = df_filt[avail].copy()
    if "Date" in df_display.columns:
        df_display["Date"] = df_display["Date"].dt.strftime("%Y-%m-%d")
    df_display = df_display.rename(columns={
        "UC_Code":   "UC Code",
        "UC_Tehsil": "Tehsil (GeoJSON)",
        "UC_No":     "UC No",
    })
    st.dataframe(df_display, use_container_width=True, height=320)

    st.divider()

    # ── UC-wise Summary Chart ─────────────────────────────────────────────
    if not df_filt.empty and df_filt["UC_Code"].notna().any():
        st.markdown('<p class="section-title">📊 Complaints by Tehsil (GeoJSON)</p>',
                    unsafe_allow_html=True)
        summary = (
            df_filt.groupby("UC_Tehsil", dropna=False)
            .size()
            .reset_index(name="Count")
            .sort_values("Count", ascending=False)
            .rename(columns={"UC_Tehsil": "Tehsil"})
        )
        st.bar_chart(summary.set_index("Tehsil")["Count"])
        st.divider()

    # ── Export Section ────────────────────────────────────────────────────
    st.markdown('<p class="section-title">📥 Export to Excel</p>',
                unsafe_allow_html=True)

    exp_info, exp_btn = st.columns([2, 1], gap="large")

    with exp_info:
        st.markdown(
            """
            The exported Excel file includes every complaint enriched with:
            - **UC Code** – from Multan District GeoJSON (spatial point-in-polygon match)
            - **Tehsil (GeoJSON)** – official Tehsil name from the boundary file
            - **UC No** – Union Council number

            Complaints whose GPS coordinates fall outside all UC polygons will
            have blank values in those columns — check the GPS field.
            """
        )
        export_all = st.checkbox("Export all data (ignore filters)", value=False)

    with exp_btn:
        export_df  = df if export_all else df_filt
        excel_buf  = generate_excel(export_df)
        ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname      = f"SPA_Multan_Complaints_{ts}.xlsx"
        st.download_button(
            label=f"⬇️  Download Excel\n({len(export_df)} records)",
            data=excel_buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
